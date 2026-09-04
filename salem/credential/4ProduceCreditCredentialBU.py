import pandas as pd
import openpyxl
import uuid

# File paths
template_path = r"C:\text\NJ\NJCCC Credit Credential Template.xlsx"
csv_path = r"C:\text\NJ\Salem\credential\parsed_credentials.csv"
output_path = r"C:\text\NJ\Salem\credential\Review\Salem_BU_Credit_Credentials.xlsx"

# Load the template workbook and CSV file
wb = openpyxl.load_workbook(template_path)
df_csv = pd.read_csv(csv_path)

# Generate CTID values for each row in the CSV file. Add this if the CSV doesn't contain CTIDs.
df_csv["CTID"] = ["ce-"+str(uuid.uuid4()) for _ in range(len(df_csv))]


# Convert the date column (assuming it's named 'date_column') to MM-DD-YYYY format
#df_csv["date_"] = pd.to_datetime(df_csv["effectiveStartDate"]).dt.strftime("%m-%d-%Y")


# Load the target worksheet (Credential Data)
ws = wb["Credential Data - MAKE UPDATES"]

# Get headers from the template (assuming headers are in row 1)
headers = [cell.value for cell in ws[1] if cell.value is not None]

# Define the column mapping from CSV to Template
column_mapping = {
    #"Credential Program Code": "Internal Identifier",
    "CTID": "CTID",
    "Name": "Credential Name",
    "Type": "Credential Type",
    "Description": "Description",
    "URL": "Subject Webpage",
    #"Hours": "ConditionProfile: Credit Unit Value",
    #"Outcomes": "Condition Profile: Required Competency Framework",
    #"CIP": "CIP List",
    "Occupation": "Occupation Type",
    "Cost": "Cost",
    #"status": "Credential Status",
    #"date_": "Date Effective",
}

# Filter and rename the CSV columns based on the mapping
df_mapped = df_csv[column_mapping.keys()].rename(columns=column_mapping)

# Add consistent values for specified columns
consistent_values = {
    "Owned By": "ce-8381aac1-e307-4e6b-83d7-3a6520e6cb6e",
    "Offered By": "ce-8381aac1-e307-4e6b-83d7-3a6520e6cb6e",
    "Credential Status": "Active",
    "Language": "English",
    "Version Identifier": "2024-2025 Catalog",
}

# Ensure all columns from the template are included
for column, value in consistent_values.items():
    if column not in df_mapped.columns:
        df_mapped[column] = value

# Ensure the order of columns matches the template headers
df_mapped = df_mapped.reindex(columns=headers, fill_value=None)

# Append data starting from row 4 in the worksheet
start_row = 4
for r_idx, row in enumerate(df_mapped.itertuples(index=False), start=4):
    for c_idx, value in enumerate(row, start=1):  # Ensure correct column offset
        ws.cell(row=r_idx, column=c_idx, value=value)

# Insert 3 additional rows with 'ADD NEW CREDENTIAL HERE' in the first column
last_row = start_row + len(df_mapped)
for i in range(3):
    ws.cell(row=last_row + i, column=1, value="ADD NEW CREDENTIAL HERE")

# Save the updated workbook
wb.save(output_path)
