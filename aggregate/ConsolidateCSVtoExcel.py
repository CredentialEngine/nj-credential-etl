"""
assemble_nj_csvs.py

Walks C:\\text\\NJ\\{institution}\\2026\\noncredit\\
and combines all CSV files into a single Excel workbook saved at:
  C:\\text\\NJ\\NJ_Noncredit_2026.xlsx

One sheet per institution; if an institution has multiple CSVs they are
stacked vertically (with a blank row between files).
"""

import glob
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR   = r"C:\text\NJ"
YEAR       = "2026"
SUB_FOLDER = "noncredit"
OUTPUT     = os.path.join(BASE_DIR, f"NJ_Noncredit_{YEAR}.xlsx")

HEADER_FILL  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT    = Font(name="Arial", size=10)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def sheet_name(institution: str) -> str:
    """Excel sheet names: max 31 chars, no special chars."""
    name = institution[:31]
    for ch in r"\/*?:[]]":
        name = name.replace(ch, "-")
    return name


def find_csvs() -> dict[str, list[str]]:
    """Return {institution: [csv_path, ...]} sorted by institution name."""
    pattern = os.path.join(BASE_DIR, "*", YEAR, SUB_FOLDER, "*.csv")
    paths = sorted(glob.glob(pattern))
    institutions: dict[str, list[str]] = {}
    for p in paths:
        parts = p.replace(BASE_DIR, "").lstrip(os.sep).split(os.sep)
        inst = parts[0]
        institutions.setdefault(inst, []).append(p)
    return dict(sorted(institutions.items()))


def format_sheet(ws, df_rows: int, df_cols: int) -> None:
    """Apply header formatting and auto-size columns."""
    for cell in ws[1]:
        cell.font    = HEADER_FONT
        cell.fill    = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=df_cols):
        for cell in row:
            cell.font = BODY_FONT

    for col_idx in range(1, df_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def main() -> None:
    institutions = find_csvs()
    if not institutions:
        print(f"No CSVs found under {BASE_DIR}. Check the path and folder structure.")
        return

    print(f"Found {sum(len(v) for v in institutions.values())} CSV(s) across "
          f"{len(institutions)} institution(s).\n")

    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        for inst, csv_paths in institutions.items():
            frames = []
            reference_cols = None
            for path in csv_paths:
                try:
                    df = pd.read_csv(path, dtype=str, encoding="utf-8-sig")
                    if reference_cols is None:
                        reference_cols = list(df.columns)
                    elif list(df.columns) != reference_cols:
                        only_in_ref  = set(reference_cols) - set(df.columns)
                        only_in_file = set(df.columns) - set(reference_cols)
                        print(f"  ⚠ Column mismatch in {os.path.basename(path)}")
                        if only_in_ref:
                            print(f"    Missing columns (will be blank): {sorted(only_in_ref)}")
                        if only_in_file:
                            print(f"    Extra columns not in first file:  {sorted(only_in_file)}")
                    frames.append(df)
                    print(f"  ✓ {os.path.basename(path)}  ({len(df)} rows)")
                except Exception as e:
                    print(f"  ✗ {os.path.basename(path)}  ERROR: {e}")

            if not frames:
                continue

            # outer join preserves all columns; missing cells become blank
            combined = pd.concat(frames, ignore_index=True, join="outer")
            sname    = sheet_name(inst)
            combined.to_excel(writer, sheet_name=sname, index=False)
            print(f"→ Sheet '{sname}': {len(combined)} total rows\n")

    # Post-process: apply formatting via openpyxl
    wb = load_workbook(OUTPUT)
    for ws in wb.worksheets:
        format_sheet(ws, ws.max_row - 1, ws.max_column)
    wb.save(OUTPUT)

    print(f"\nDone! Saved to: {OUTPUT}")


if __name__ == "__main__":
    main()