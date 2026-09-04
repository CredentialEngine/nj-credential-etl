import pandas as pd
import openpyxl
import uuid

# File paths
template_path = r"C:\text\NJ\NJCCC Credit Credential Template.xlsx"
csv_path = r"C:\text\NJ\Bergen\2024-2025\Review\Bergen_BU_Credit_Credentials.csv"
output_path = r"C:\text\NJ\Bergen\2024-2025\Review\Bergen_BU_Credit_Credentials.xlsx"

# Load the template workbook and CSV file
wb = openpyxl.load_workbook(template_path)
df_csv = pd.read_csv(csv_path)

# Generate CTID values for each row in the CSV file. Add this if the CSV doesn't contain CTIDs.
#df_csv["CTID"] = ["ce-"+str(uuid.uuid4()) for _ in range(len(df_csv))]

# Load the target worksheet (Credential Data)
ws = wb["Credential Data - MAKE UPDATES"]

# Get headers from the template (assuming headers are in row 1)
headers = [cell.value for cell in ws[1] if cell.value is not None]

# Define the column mapping from CSV to Template
column_mapping = {
    "Code": "Internal Identifier",
    "CTID": "CTID",
    "Program Name": "Credential Name",
    "Credential Type": "Credential Type",
    "Description": "Description",
    "Subject Webpage": "Subject Webpage",
    "Credit": "ConditionProfile: Credit Unit Value",
    "Condition Profile: Required Competency Framework": "Condition Profile: Required Competency Framework",
}

# Filter and rename the CSV columns based on the mapping
df_mapped = df_csv[column_mapping.keys()].rename(columns=column_mapping)

# Add consistent values for specified columns
consistent_values = {
    "Owned By": "ce-84b7d711-b47e-412b-a89c-2e8165db56b2",
    "Offered By": "ce-84b7d711-b47e-412b-a89c-2e8165db56b2",
    "Credential Status": "Active",
    "Language": "English",
    "Version Identifier": "2024-2025 Catalog",
}

# Ensure all columns from the template are included
for column, value in consistent_values.items():
    if column not in df_mapped.columns:
        df_mapped[column] = value

# Ensure the order of columns matches the template headers
df_mapped = df_mapped.reindex(columns=headers, fill_value=None)

# Append data starting from row 4 in the worksheet
start_row = 4
for r_idx, row in enumerate(df_mapped.itertuples(index=False), start=4):
    for c_idx, value in enumerate(row, start=1):  # Ensure correct column offset
        ws.cell(row=r_idx, column=c_idx, value=value)

# Insert 3 additional rows with 'ADD NEW CREDENTIAL HERE' in the first column
last_row = start_row + len(df_mapped)
for i in range(3):
    ws.cell(row=last_row + i, column=1, value="ADD NEW CREDENTIAL HERE")

# Save the updated workbook
wb.save(output_path)
