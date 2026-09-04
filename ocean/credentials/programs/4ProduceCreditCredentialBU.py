import pandas as pd
import openpyxl
import uuid

# File paths
template_path = r"C:\text\NJ\NJCCC Credit Credential Template.xlsx"
csv_path = r"C:\text\NJ\Ocean\credentials\programs\parsed_credentials_updated.csv"
output_path = r"C:\text\NJ\Ocean\credentials\programs\Review\Ocean_BU_Credit_Credentials.xlsx"

# Load the template workbook and CSV file
wb = openpyxl.load_workbook(template_path)
df_csv = pd.read_csv(csv_path)

# Generate CTID values for each row in the CSV file. Add this if the CSV doesn't contain CTIDs.
df_csv["CTID"] = ["ce-"+str(uuid.uuid4()) for _ in range(len(df_csv))]

# Create the 'Subject Webpage' column by combining text and 'Program Code'
#df_csv["Subject Webpage"] = "https://catalog.brookdalecc.edu/programs/" + df_csv["Program Code"] 

# Convert the date column (assuming it's named 'date_column') to MM-DD-YYYY format
#df_csv["date_"] = pd.to_datetime(df_csv["effectiveStartDate"]).dt.strftime("%m-%d-%Y")


# Load the target worksheet (Credential Data)
ws = wb["Credential Data - MAKE UPDATES"]

# Get headers from the template (assuming headers are in row 1)
headers = [cell.value for cell in ws[1] if cell.value is not None]

# Define the column mapping from CSV to Template
column_mapping = {
    #"Credential Program Code": "Internal Identifier",
    "CTID": "CTID",
    "Name": "Credential Name",
    "Type": "Credential Type",
    "Description": "Description",
    "URL": "Subject Webpage",
    "Hours": "ConditionProfile: Credit Unit Value",
    "Cost": "Cost",
    "Semesters": "Estimated Duration"
    #"Outcomes": "Condition Profile: Required Competency Framework",
    #"CIP": "CIP List",
    #"Occupation": "Occupation Type",
    #"status": "Credential Status",
    #"date_": "Date Effective",
}

# Filter and rename the CSV columns based on the mapping
df_mapped = df_csv[column_mapping.keys()].rename(columns=column_mapping)

# Add consistent values for specified columns
consistent_values = {
    "Owned By": "ce-6d4a0b3c-45d0-49ff-abe9-a83e587581e5",
    "Offered By": "ce-6d4a0b3c-45d0-49ff-abe9-a83e587581e5",
    "Credential Status": "Active",
    "Language": "English",
    "Version Identifier": "2024-2025 Catalog",
}

# Ensure all columns from the template are included
for column, value in consistent_values.items():
    if column not in df_mapped.columns:
        df_mapped[column] = value

# Ensure the order of columns matches the template headers
df_mapped = df_mapped.reindex(columns=headers, fill_value=None)

# Append data starting from row 4 in the worksheet
start_row = 4
for r_idx, row in enumerate(df_mapped.itertuples(index=False), start=4):
    for c_idx, value in enumerate(row, start=1):  # Ensure correct column offset
        ws.cell(row=r_idx, column=c_idx, value=value)

# Insert 3 additional rows with 'ADD NEW CREDENTIAL HERE' in the first column
last_row = start_row + len(df_mapped)
for i in range(3):
    ws.cell(row=last_row + i, column=1, value="ADD NEW CREDENTIAL HERE")

# Save the updated workbook
wb.save(output_path)
